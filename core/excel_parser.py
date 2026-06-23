#!/usr/bin/env python3
"""Trình phân tích tệp Excel cho thao tác nhập hàng."""
import openpyxl
from io import BytesIO
from typing import List, Dict, Optional
from core.logger import get_logger

logger = get_logger(__name__)


class ExcelParser:
    """Phân tích tệp Excel cho các giao dịch nhập hàng."""

    @staticmethod
    def parse_import_file(file_content: bytes, sheet_name: str = None) -> Dict:
        """
        Phân tích tệp Excel chứa dữ liệu nhập hàng.

        Định dạng kỳ vọng:
        - Cột A: Mã sản phẩm (hoặc ID sản phẩm)
        - Cột B: Tên sản phẩm (tùy chọn, sẽ tra cứu nếu cung cấp mã)
        - Cột C: Số lượng
        - Cột D: Đơn giá

        Trả về: {
            'success': bool,
            'items': List[{product_code, quantity, unit_price}],
            'error': str (nếu không thành công)
        }
        """
        try:
            # Tải workbook từ bytes
            wb = openpyxl.load_workbook(BytesIO(file_content))

            # Lấy sheet (sheet đầu tiên theo mặc định hoặc sheet cụ thể)
            if sheet_name:
                if sheet_name not in wb.sheetnames:
                    return {
                        'success': False,
                        'error': f"Không tìm thấy sheet '{sheet_name}'. Có sẵn: {', '.join(wb.sheetnames)}"
                    }
                ws = wb[sheet_name]
            else:
                ws = wb.active

            logger.info(f"Đang phân tích sheet Excel: {ws.title}")

            items = []
            row_count = 0
            error_rows = []

            # Bỏ qua hàng tiêu đề (hàng 1)
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
                row_count += 1

                try:
                    # Trích xuất giá trị
                    product_code = row[0].value if row[0] else None
                    product_name = row[1].value if len(row) > 1 else None
                    quantity = row[2].value if len(row) > 2 else None
                    unit_price = row[3].value if len(row) > 3 else None

                    # Kiểm tra các trường bắt buộc
                    if not product_code:
                        continue  # Bỏ qua hàng trống

                    if not quantity or not unit_price:
                        error_rows.append(f"Hàng {row_idx}: Thiếu số lượng hoặc đơn giá")
                        continue

                    try:
                        qty = float(quantity)
                        price = float(unit_price)
                    except (ValueError, TypeError):
                        error_rows.append(f"Hàng {row_idx}: Định dạng số lượng/đơn giá không hợp lệ")
                        continue

                    items.append({
                        'product_code': str(product_code).strip(),
                        'product_name': str(product_name).strip() if product_name else None,
                        'quantity': qty,
                        'unit_price': price
                    })

                except Exception as e:
                    logger.warning(f"Lỗi khi phân tích hàng {row_idx}: {e}")
                    error_rows.append(f"Hàng {row_idx}: {str(e)}")
                    continue

            if not items:
                return {
                    'success': False,
                    'error': f"Không tìm thấy mục hợp lệ nào trong tệp Excel. Lỗi: {'; '.join(error_rows) if error_rows else 'Tệp rỗng'}"
                }

            result = {
                'success': True,
                'items': items,
                'row_count': len(items)
            }

            if error_rows:
                result['warnings'] = error_rows

            logger.info(f"Đã phân tích thành công {len(items)} mục từ Excel")
            return result

        except openpyxl.utils.exceptions.InvalidFileException:
            return {
                'success': False,
                'error': 'Định dạng tệp Excel không hợp lệ. Vui lòng tải lên tệp .xlsx.'
            }
        except Exception as e:
            logger.error(f"Lỗi khi phân tích tệp Excel: {e}", exc_info=True)
            return {
                'success': False,
                'error': f"Lỗi khi phân tích tệp Excel: {str(e)}"
            }

    @staticmethod
    def get_sample_template() -> bytes:
        """Sinh mẫu Excel để nhập hàng."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Mẫu nhập hàng"

        # Tiêu đề cột
        headers = ["Mã sản phẩm", "Tên sản phẩm", "Số lượng", "Đơn giá"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        
        # Dữ liệu mẫu
        sample_data = [
            ["SKU001", "Sản phẩm A", 10, 50000],
            ["SKU002", "Sản phẩm B", 5, 75000],
        ]
        for row_idx, row_data in enumerate(sample_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx).value = value

        # Điều chỉnh độ rộng cột
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15
        
        # Chuyển sang bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
