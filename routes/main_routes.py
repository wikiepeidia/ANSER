"""Core product/customer API and logout — main_bp Blueprint."""
import json
import os
import tempfile
import uuid

from flask import Blueprint, flash, jsonify, redirect, request, session, url_for
from flask_login import current_user, login_required, logout_user
from core.config import Config
from core.extensions import db_manager
from core.logger import get_logger
from core.security import safe_api_error, validate_upload

from core.services.customer_service import (
    create_customer, delete_customer, get_all_customers, update_customer,
)
from core.services.product_service import (
    create_product, delete_product, detect_column_mapping, get_all_products,
    import_products_from_excel, update_product, FIELD_LABELS,
)

main_bp = Blueprint('main', __name__)
logger = get_logger(__name__)

# Temp file cache: {uuid: path} — file lưu sau preview, dùng lại khi import
_import_temp_files: dict = {}

EXCEL_EXTENSIONS = {'xlsx', 'xls'}
EXCEL_MIMETYPES = {
    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    'application/vnd.ms-excel',
}


@main_bp.route('/logout')
@login_required
def logout():
    # Clears the shared session cookie — logs the user out of every mảng,
    # since Gateway (:5000), Retail (:5002) and Sản xuất (:5003) all trust
    # the same cookie.
    logout_user()
    flash('Bạn đã đăng xuất thành công.', 'success')
    return redirect(f"{Config.GATEWAY_ORIGIN}/auth/signin")


# ── Customers ──────────────────────────────────────────────────────────────

@main_bp.route('/api/customers', methods=['GET'])
@login_required
def api_get_customers():
    conn = db_manager.get_business_connection()
    try:
        customers = get_all_customers(conn, current_user.id, getattr(current_user, 'role', 'user'))
        return jsonify({'success': True, 'customers': customers})
    finally:
        conn.close()


@main_bp.route('/api/customers', methods=['POST'])
@login_required
def api_create_customer():
    data = request.get_json()
    if not data or 'code' not in data or 'name' not in data:
        return jsonify({'success': False, 'message': 'Thiếu các trường bắt buộc'}), 400
    conn = db_manager.get_business_connection()
    try:
        ok, err = create_customer(
            conn, data['code'], data['name'], data.get('phone', ''), data.get('email', ''),
            data.get('address', ''), data.get('notes', ''), current_user.id,
        )
        if ok:
            return jsonify({'success': True, 'message': 'Tạo khách hàng thành công'})
        return jsonify({'success': False, 'message': err}), 400
    finally:
        conn.close()


@main_bp.route('/api/customers/<int:customer_id>', methods=['PUT'])
@login_required
def api_update_customer(customer_id):
    data = request.get_json()
    conn = db_manager.get_business_connection()
    try:
        update_customer(
            conn, customer_id, data['name'], data.get('phone', ''), data.get('email', ''),
            data.get('address', ''), data.get('notes', ''), current_user.id,
            getattr(current_user, 'role', 'user'),
        )
        return jsonify({'success': True, 'message': 'Cập nhật khách hàng thành công'})
    except LookupError as e:
        return jsonify({'success': False, 'message': str(e)}), 404
    finally:
        conn.close()


@main_bp.route('/api/customers/<int:customer_id>', methods=['DELETE'])
@login_required
def api_delete_customer(customer_id):
    conn = db_manager.get_business_connection()
    try:
        delete_customer(conn, customer_id, current_user.id, getattr(current_user, 'role', 'user'))
        return jsonify({'success': True, 'message': 'Xóa khách hàng thành công'})
    except LookupError as e:
        return jsonify({'success': False, 'message': str(e)}), 404
    finally:
        conn.close()


# ── Products ───────────────────────────────────────────────────────────────

@main_bp.route('/api/products', methods=['GET'])
@login_required
def api_get_products():
    conn = db_manager.get_business_connection()
    try:
        products = get_all_products(
            conn, session.get('active_warehouse_id'), current_user.id,
            getattr(current_user, 'role', 'user'),
        )
        return jsonify({'success': True, 'products': products})
    finally:
        conn.close()


@main_bp.route('/api/products', methods=['POST'])
@login_required
def api_create_product():
    data = request.get_json()
    if not data or 'code' not in data or 'name' not in data:
        return jsonify({'success': False, 'message': 'Thiếu các trường bắt buộc'}), 400
    conn = db_manager.get_business_connection()
    try:
        ok, err = create_product(
            conn, data['code'], data['name'], data.get('category', ''), data.get('unit', 'cái'),
            data.get('price', 0), data.get('stock_quantity', 0), data.get('description', ''),
            current_user.id, data.get('image_url', ''),
        )
        if ok:
            return jsonify({'success': True, 'message': 'Tạo sản phẩm thành công'})
        return jsonify({'success': False, 'message': err}), 400
    finally:
        conn.close()


@main_bp.route('/api/products/<int:product_id>', methods=['PUT'])
@login_required
def api_update_product(product_id):
    data = request.get_json()
    conn = db_manager.get_business_connection()
    try:
        update_product(
            conn, product_id, data['name'], data.get('category', ''), data.get('unit', 'cái'),
            data.get('price', 0), data.get('stock_quantity', 0), data.get('description', ''),
            data.get('image_url', ''), current_user.id, getattr(current_user, 'role', 'user'),
        )
        return jsonify({'success': True, 'message': 'Cập nhật sản phẩm thành công'})
    except LookupError as e:
        return jsonify({'success': False, 'message': str(e)}), 404
    finally:
        conn.close()


@main_bp.route('/api/products/<int:product_id>', methods=['DELETE'])
@login_required
def api_delete_product(product_id):
    conn = db_manager.get_business_connection()
    try:
        delete_product(conn, product_id, current_user.id, getattr(current_user, 'role', 'user'))
        return jsonify({'success': True, 'message': 'Xóa sản phẩm thành công'})
    except LookupError as e:
        return jsonify({'success': False, 'message': str(e)}), 404
    finally:
        conn.close()


@main_bp.route('/api/products/import-excel/preview', methods=['POST'])
@login_required
def api_import_products_excel_preview():
    """Đọc header, lưu file vào temp, trả về mapping + file_id.

    Import step sau đó chỉ cần gửi file_id — không upload lại file.
    """
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': 'Không có file'}), 400
    file = request.files['file']
    try:
        validate_upload(file, EXCEL_EXTENSIONS, EXCEL_MIMETYPES)
    except ValueError as e:
        return jsonify({'success': False, 'message': str(e)}), 400
    try:
        from io import BytesIO
        from openpyxl import load_workbook

        file_bytes = file.read()

        # Lưu file vào temp để import step dùng lại (tránh upload 2 lần)
        suffix = '.xlsx' if file.filename.lower().endswith('.xlsx') else '.xls'
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
        tmp.write(file_bytes)
        tmp.close()
        file_id = str(uuid.uuid4())
        _import_temp_files[file_id] = tmp.name

        # Đọc header + 2 dòng mẫu bằng streaming parser
        wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)

        raw_cols = [str(h).strip() if h is not None else '' for h in next(rows_iter)]
        raw_cols = [c for c in raw_cols if c]

        sample_rows = []
        for row in rows_iter:
            if len(sample_rows) >= 2:
                break
            sample_rows.append(row)
        wb.close()

        mapping = detect_column_mapping(raw_cols)
        for item in mapping:
            col_idx = raw_cols.index(item['original'])
            samples = []
            for row in sample_rows:
                if col_idx < len(row) and row[col_idx] is not None:
                    val = str(row[col_idx]).strip()
                    if val and val != 'None':
                        samples.append(val)
            item['samples'] = samples[:2]

        return jsonify({
            'success': True,
            'mapping': mapping,
            'field_labels': FIELD_LABELS,
            'file_id': file_id,
        })
    except Exception as e:
        return safe_api_error(logger, message='Không thể đọc file Excel', status=400, exc=e)


@main_bp.route('/api/products/import-excel', methods=['POST'])
@login_required
def api_import_products_excel():
    column_map = None
    raw_map = request.form.get('column_map')
    if raw_map:
        try:
            column_map = json.loads(raw_map)
        except (ValueError, TypeError):
            pass

    # Dùng temp file từ preview nếu có — tránh upload lại
    file_id = request.form.get('file_id')
    file_source = None

    if file_id and file_id in _import_temp_files:
        tmp_path = _import_temp_files.pop(file_id)
        try:
            file_source = open(tmp_path, 'rb')
        except OSError:
            file_source = None
        finally:
            try:
                os.unlink(tmp_path)
            except OSError:
                pass

    if file_source is None:
        if 'file' not in request.files or not request.files['file'].filename:
            return jsonify({'success': False, 'message': 'Không có file'}), 400
        f = request.files['file']
        try:
            validate_upload(f, EXCEL_EXTENSIONS, EXCEL_MIMETYPES)
        except ValueError as e:
            return jsonify({'success': False, 'message': str(e)}), 400
        file_source = f

    conn = db_manager.get_business_connection()
    try:
        result = import_products_from_excel(conn, file_source, current_user.id, column_map=column_map)
        return jsonify({
            'success': True,
            'message': f"Import thành công: {result['inserted']} thêm mới, {result['updated']} cập nhật, {result['skipped']} bỏ qua",
            **result,
        })
    except ValueError as e:
        return jsonify({'success': False, 'message': str(e)}), 400
    except Exception as e:
        return safe_api_error(logger, exc=e)
    finally:
        conn.close()
        if hasattr(file_source, 'close'):
            file_source.close()


# ── Active warehouse selection (cửa hàng đang thao tác) ─────────────────────
# Bất kỳ role nào cũng cần chọn được (nhân viên đứng ở cửa hàng nào thì thao
# tác cho cửa hàng đó) — khác với /api/admin/warehouses (CRUD, admin-only).

@main_bp.route('/api/warehouses', methods=['GET'])
@login_required
def api_list_warehouses_for_switcher():
    conn = db_manager.get_business_connection()
    try:
        c = conn.cursor()
        c.execute('SELECT id, name FROM warehouses WHERE is_active = 1 ORDER BY name')
        warehouses = [{'id': r['id'], 'name': r['name']} for r in c.fetchall()]
        return jsonify({
            'success': True,
            'warehouses': warehouses,
            'active_warehouse_id': session.get('active_warehouse_id'),
        })
    finally:
        conn.close()


@main_bp.route('/api/session/active-warehouse', methods=['POST'])
@login_required
def api_set_active_warehouse():
    data = request.get_json(silent=True) or {}
    warehouse_id = data.get('warehouse_id')
    if warehouse_id in (None, '', 'null'):
        session.pop('active_warehouse_id', None)
        return jsonify({'success': True, 'active_warehouse_id': None})
    session['active_warehouse_id'] = int(warehouse_id)
    return jsonify({'success': True, 'active_warehouse_id': session['active_warehouse_id']})
