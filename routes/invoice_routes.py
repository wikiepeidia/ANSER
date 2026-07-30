"""Invoice import + file attachment routes (INVOICE-01, INVOICE-02).

INVOICE-01: POST /api/invoices/import reads an uploaded .xlsx workbook and
creates one material_batches row per valid data row, reusing the exact same
supplier/material resolution helpers routes/inventory_routes.py already
established (_resolve_supplier, _resolve_material_product, _find_material) —
per 04-CONTEXT.md's explicit instruction not to reimplement that logic. A
row with a missing/invalid materialCode or a non-positive quantity is
skipped and reported by row number + reason; the whole file only fails a
missing file or a non-.xlsx filename.

INVOICE-02: POST/GET /api/invoices/attachments upload/retrieve/list invoice
file attachments (format allowlist pdf/jpg/png/xlsx, max 10MB), backed by
the invoice_attachments table 04-02 created. Format and size are validated
BEFORE any disk write or DB insert (T-04-05).
"""
import os
import secrets

from flask import Blueprint, jsonify, request, send_from_directory
from flask_login import current_user, login_required
from openpyxl import load_workbook
from werkzeug.utils import secure_filename

from core.sanxuat_db import get_connection, now
from routes.inventory_routes import _find_material, _resolve_material_product, _resolve_supplier

invoice_bp = Blueprint('invoice', __name__)

# Canonical field -> accepted header strings (already lowercased/stripped
# when compared against a parsed header cell). Header matching tolerates
# case and surrounding whitespace per 04-03-PLAN.md's behavior spec.
_HEADER_ALIASES = {
    'materialCode': {'mã nvl', 'ma nvl', 'materialcode', 'material_code', 'mã nguyên vật liệu'},
    'quantity': {'số lượng', 'so luong', 'quantity'},
    'supplierName': {'nhà cung cấp', 'nha cung cap', 'suppliername', 'supplier_name', 'supplier'},
    'expiryDate': {'hạn sử dụng', 'han su dung', 'expirydate', 'expiry_date', 'expiry'},
    'notes': {'ghi chú', 'ghi chu', 'notes', 'note'},
}

# Invoice attachments (INVOICE-02): hard format allowlist + size limit,
# enforced BEFORE any disk write or DB insert (T-04-05/T-04-04).
ALLOWED_ATTACHMENT_EXTENSIONS = {'pdf', 'jpg', 'png', 'xlsx'}
MAX_ATTACHMENT_SIZE_BYTES = 10 * 1024 * 1024
UPLOAD_DIR = os.path.join('uploads', 'invoices')


# ── Excel invoice/material import (INVOICE-01) ──────────────────────────

@invoice_bp.route('/api/invoices/import', methods=['POST'])
@login_required
def import_invoice_excel():
    file = request.files.get('file')
    if not file or not (file.filename or '').lower().endswith('.xlsx'):
        return jsonify({
            'success': False,
            'message': 'Thiếu file hoặc file không đúng định dạng .xlsx',
        }), 400

    wb = load_workbook(file, data_only=True)
    ws = wb.active

    header_map = {}
    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    for i, cell in enumerate(header_row):
        normalized = str(cell.value or '').strip().lower()
        for field, aliases in _HEADER_ALIASES.items():
            if normalized in aliases:
                header_map[field] = i

    if 'materialCode' not in header_map or 'quantity' not in header_map:
        return jsonify({
            'success': False,
            'message': 'File thiếu cột bắt buộc (Mã NVL, Số lượng)',
        }), 400

    conn = get_connection()
    imported = []
    errors = []
    try:
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if all(cell is None for cell in row):
                continue

            material_code_raw = row[header_map['materialCode']]
            material = _find_material(str(material_code_raw).strip()) if material_code_raw else None
            if not material:
                errors.append({'row': row_idx, 'reason': 'Thiếu hoặc sai mã nguyên vật liệu'})
                continue

            quantity_raw = row[header_map['quantity']]
            try:
                quantity = float(quantity_raw)
            except (TypeError, ValueError):
                quantity = 0
            if quantity <= 0:
                errors.append({'row': row_idx, 'reason': 'Số lượng phải lớn hơn 0'})
                continue

            supplier_name = ''
            if 'supplierName' in header_map and row[header_map['supplierName']]:
                supplier_name = str(row[header_map['supplierName']]).strip()
            expiry_date = ''
            if 'expiryDate' in header_map and row[header_map['expiryDate']]:
                expiry_date = str(row[header_map['expiryDate']]).strip()
            notes = ''
            if 'notes' in header_map and row[header_map['notes']]:
                notes = str(row[header_map['notes']]).strip()

            try:
                supplier_id = _resolve_supplier(conn, supplier_name, current_user.id)
                material_product_id = _resolve_material_product(conn, material, current_user.id)
                cur = conn.execute(
                    'INSERT INTO material_batches (material_code, material_name, material_product_id, '
                    'unit, supplier_id, quantity, expiry_date, notes, created_by, created_at) '
                    'VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)',
                    (material['code'], material['name'], material_product_id, material['unit'],
                     supplier_id, quantity, expiry_date, notes, current_user.id, now()),
                )
                batch_id = cur.lastrowid
                code = f'LO-{2000 + batch_id}'
                conn.execute('UPDATE material_batches SET batch_code = ? WHERE id = ?', (code, batch_id))
                imported.append({'row': row_idx, 'id': batch_id, 'code': code})
            except Exception as e:
                errors.append({'row': row_idx, 'reason': str(e)})

        conn.commit()
        return jsonify({
            'success': True,
            'importedCount': len(imported),
            'errorCount': len(errors),
            'imported': imported,
            'errors': errors,
        })
    finally:
        conn.close()


# ── Invoice file attachments (INVOICE-02) ────────────────────────────────

@invoice_bp.route('/api/invoices/attachments', methods=['POST'])
@login_required
def upload_invoice_attachment():
    file = request.files.get('file')
    if not file or not file.filename:
        return jsonify({'success': False, 'message': 'Thiếu file đính kèm'}), 400

    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
    if ext not in ALLOWED_ATTACHMENT_EXTENSIONS:
        return jsonify({
            'success': False,
            'message': 'Định dạng file không được hỗ trợ (chỉ nhận pdf/jpg/png/xlsx)',
        }), 400

    file.stream.seek(0, os.SEEK_END)
    size_bytes = file.stream.tell()
    file.stream.seek(0)
    if size_bytes > MAX_ATTACHMENT_SIZE_BYTES:
        return jsonify({
            'success': False,
            'message': 'File vượt quá dung lượng tối đa 10MB',
        }), 400

    ref_type = request.form.get('refType') or 'material_batch'
    ref_id = request.form.get('refId', type=int)

    os.makedirs(UPLOAD_DIR, exist_ok=True)
    stored_name = f'{secrets.token_hex(8)}_{secure_filename(file.filename)}'
    stored_path = os.path.join(UPLOAD_DIR, stored_name)
    file.save(stored_path)

    conn = get_connection()
    try:
        cur = conn.execute(
            'INSERT INTO invoice_attachments (ref_type, ref_id, file_name, file_path, '
            'content_type, size_bytes, uploaded_at, created_by) VALUES (?, ?, ?, ?, ?, ?, ?, ?)',
            (ref_type, ref_id, file.filename, stored_path, file.content_type, size_bytes,
             now(), current_user.id),
        )
        attachment_id = cur.lastrowid
        conn.commit()
        return jsonify({
            'success': True,
            'id': attachment_id,
            'fileName': file.filename,
            'filePath': stored_path,
        })
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'message': str(e)}), 400
    finally:
        conn.close()


@invoice_bp.route('/api/invoices/attachments/<int:attachment_id>', methods=['GET'])
@login_required
def download_invoice_attachment(attachment_id):
    conn = get_connection()
    try:
        row = conn.execute(
            'SELECT * FROM invoice_attachments WHERE id = ?', (attachment_id,)
        ).fetchone()
    finally:
        conn.close()
    if not row:
        return jsonify({'success': False, 'message': 'Không tìm thấy file đính kèm'}), 404
    directory, filename = os.path.split(row['file_path'])
    return send_from_directory(directory, filename, as_attachment=True, download_name=row['file_name'])


@invoice_bp.route('/api/invoices/attachments', methods=['GET'])
@login_required
def list_invoice_attachments():
    ref_type = request.args.get('refType')
    ref_id = request.args.get('refId', type=int)

    conn = get_connection()
    try:
        if ref_type and ref_id is not None:
            rows = conn.execute(
                'SELECT * FROM invoice_attachments WHERE ref_type = ? AND ref_id = ? '
                'ORDER BY uploaded_at DESC',
                (ref_type, ref_id),
            ).fetchall()
        else:
            rows = conn.execute(
                'SELECT * FROM invoice_attachments ORDER BY uploaded_at DESC'
            ).fetchall()
    finally:
        conn.close()

    return jsonify({
        'success': True,
        'attachments': [
            {
                'id': r['id'],
                'refType': r['ref_type'],
                'refId': r['ref_id'],
                'fileName': r['file_name'],
                'contentType': r['content_type'],
                'sizeBytes': r['size_bytes'],
                'uploadedAt': r['uploaded_at'],
            }
            for r in rows
        ],
    })
